from openpyxl import load_workbook


class ExcelReader:

    REQUIRED_COLUMNS = [
        "NameA",
        "NameE",
        "Email"
    ]

    def __init__(self, file_path):
        self.file_path = file_path

    def read(self):

        workbook = load_workbook(self.file_path, data_only=True)
        sheet = workbook.active

        headers = {}

        for index, cell in enumerate(sheet[1], start=1):
            headers[str(cell.value).strip()] = index

        for column in self.REQUIRED_COLUMNS:
            if column not in headers:
                raise Exception(f"العمود {column} غير موجود داخل ملف Excel")

        participants = []

        for row in range(2, sheet.max_row + 1):

            name_ar = sheet.cell(row=row,
                                 column=headers["NameA"]).value

            name_en = sheet.cell(row=row,
                                 column=headers["NameE"]).value

            email = sheet.cell(row=row,
                               column=headers["Email"]).value

            if not name_ar:
                continue

            participants.append({

                "name_ar": str(name_ar).strip(),

                "name_en": "" if name_en is None else str(name_en).strip(),

                "email": "" if email is None else str(email).strip()

            })

        workbook.close()

        return participants
